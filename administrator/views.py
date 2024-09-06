from django.shortcuts import render,redirect
from website.models import Custumer, Kategori, Kontak, Profil, Bahan, Slide, Testimoni
from cart.models import Transaksi, DetailTransaksi
from .forms import KategoriForm,BahanForm,SlideForm,TestimoniForm
from django.contrib.auth.decorators import login_required
from website.decorators import ijinkan_pengguna,pilihan_login
from django.db.models import Count, Q, Sum
from django.contrib.auth.models import User
from django.contrib.auth.models import Group

import json

import datetime

from datetime import date
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment



@login_required(login_url='loginpage')
@pilihan_login
def beranda_admin(request):
    jmlKategori = Kategori.objects.filter(aktif=True).count()
    jmlCustumer = Custumer.objects.count()
    jmlBahan = Bahan.objects.count()
    jmlTransaksi = Transaksi.objects.filter(status="Lunas").count()
    # grafik = Transaksi.objects.annotate(bulan=Count('date_created__month')).values('bulan')

    tahun_ini = datetime.now().year
    bulan_ini = datetime.now().month

    daftar_bulan = [
        {'id': i, 'nama': get_nama_bulan(i)} 
        for i in range(1, bulan_ini + 1)
    ]
   

    data_transaksi = Transaksi.objects.filter(date_created__year=tahun_ini, date_created__month__lte=bulan_ini, status='Lunas')\
        .values('date_created__month')\
        .annotate(jumlah=Count('id'))
    
    data_transaksi_per_bulan = [0] * bulan_ini
    for data in data_transaksi:
        bulan = data['date_created__month']
        jumlah = data['jumlah']
        data_transaksi_per_bulan[bulan-1] = jumlah

    context = {
        'judul': 'Halaman Beranda',
        'menu': 'beranda',
        'jmlKategori':jmlKategori,
        'jmlCustumer':jmlCustumer,
        'jmlBahan':jmlBahan,
        'jmlTransaksi':jmlTransaksi,
        'daftar_bulan':json.dumps(daftar_bulan),
        'data_transaksi':json.dumps(data_transaksi_per_bulan),
        'tahun_ini':tahun_ini,


    }
    return render(request, 'admin_beranda.html', context)
def get_nama_bulan(bulan):
    nama_bulan = {
        1: 'Januari',
        2: 'Februari',
        3: 'Maret',
        4: 'April',
        5: 'Mei',
        6: 'Juni',
        7: 'Juli',
        8: 'Agustus',
         9: 'September',
        10: 'Oktober',
        11: 'November',
        12: 'Desember'
    }
    return nama_bulan.get(bulan, '')

@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def kategori_admin(request):
    kategori = Kategori.objects.all()
    context = {
        'data': kategori,
        'judul': 'Halaman Kategori',
        'menu': 'kategori'
    }
    return render(request, 'admin_kategori.html', context)

@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def formkategori_admin(request):
    form = KategoriForm()
    if request.method == 'POST':
        formsimpan = KategoriForm(request.POST)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('kategori_admin')
    context = {
        'judul': 'Form Kategori',
         'menu': 'kategori',
        'form':form
    }
    return render(request, 'admin_formkategori.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def editkategori_admin(request, slug):
    kategori = Kategori.objects.get(slug=slug)
    form = KategoriForm(instance=kategori)
    if request.method == 'POST':
        formsimpan = KategoriForm(request.POST, instance=kategori)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('kategori_admin')
    context = {
        'judul': 'Form Edit Kategori',
        'form':form,
        'menu': 'kategori',
    }
    return render(request, 'admin_formkategori.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def deletekategori_admin(request, pk):
    hapus = Kategori.objects.get(id=pk)
    if request.method == 'POST':
        hapus.delete()
        return redirect('kategori_admin')

    context = {
        'judul': 'Form Hapus Kategori',
        'hapus':hapus,
        'menu': 'kategori',
    }
    return render(request, 'admin_hapuskategori.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def bahan_admin(request):
    bahan = Bahan.objects.all()
    context = {
        'data': bahan,
        'judul': 'Halaman Bahan Meubel',
        'menu': 'bahan'
    }
    return render(request, 'admin_bahan.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def formbahan_admin(request):
    form = BahanForm()
    if request.method == 'POST':
        formsimpan = BahanForm(request.POST, request.FILES)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('bahan_admin')
    context = {
        'judul': 'Form Bahan',
         'menu': 'bahan',
        'form':form
    }
    return render(request, 'admin_formbahan.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def editbahan_admin(request, slug):
    bahan = Bahan.objects.get(slug=slug)
    form = BahanForm(instance=bahan)
    if request.method == 'POST':
        formsimpan = BahanForm(request.POST,request.FILES, instance=bahan)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('bahan_admin')
    context = {
        'judul': 'Form Edit Bahan',
        'form':form,
        'menu': 'bahan',
    }
    return render(request, 'admin_formbahan.html', context)



@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def deletebahan_admin(request, pk):
    hapus = Bahan.objects.get(id=pk)
    if request.method == 'POST':
        hapus.delete()
        return redirect('bahan_admin')
    context = {
        'judul': 'Form Hapus Bahan',
        'hapus':hapus,
        'menu': 'bahan',
    }
    return render(request, 'admin_hapusbahan.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def slide_admin(request):
    slide = Slide.objects.all()
    context = {
        'data': slide,
        'judul': 'Halaman Slide',
        'menu': 'slide'
    }
    return render(request, 'admin_slide.html', context)



@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def formslide_admin(request):
    form = SlideForm()
    if request.method == 'POST':
        formsimpan = SlideForm(request.POST,request.FILES)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('slide_admin')
    context = {
        'judul': 'Form Slide',
         'menu': 'slide',
        'form':form
    }
    return render(request, 'admin_formslide.html', context)
@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def editslide_admin(request, pk):
    slide = Slide.objects.get(id=pk)
    form = SlideForm(instance=slide)
    if request.method == 'POST':
        formsimpan = SlideForm(request.POST,request.FILES, instance=slide)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('slide_admin')
    context = {
        'judul': 'Form Edit Slide',
        'form':form,
        'menu': 'slide',
    }
    return render(request, 'admin_formslide.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def deleteslide_admin(request, pk):
    hapus = Slide.objects.get(id=pk)
    if request.method == 'POST':
        hapus.delete()
        return redirect('slide_admin')

    context = {
        'judul': 'Form Hapus Slide',
        'hapus':hapus,
        'menu': 'slide',
    }
    return render(request, 'admin_hapusslide.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def testimoni_admin(request):
    testimoni = Testimoni.objects.all()
    context = {
        'data': testimoni,
        'judul': 'Halaman Testimoni',
        'menu': 'testimoni'
    }
    return render(request, 'admin_testimoni.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def formtestimoni_admin(request):
    form = TestimoniForm()
    if request.method == 'POST':
        formsimpan = TestimoniForm(request.POST,request.FILES)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('testimoni_admin')
    context = {
        'judul': 'Form Testimoni',
         'menu': 'testimoni',
        'form':form
    }
    return render(request, 'admin_formtestimoni.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def edittestimoni_admin(request, pk):
    testimoni = Testimoni.objects.get(id=pk)
    form = TestimoniForm(instance=testimoni)
    if request.method == 'POST':
        formsimpan = TestimoniForm(request.POST,request.FILES, instance=testimoni)
        if formsimpan.is_valid():
            formsimpan.save()
            return redirect('testimoni_admin')
    context = {
        'judul': 'Form Edit Testimoni',
        'form':form,
        'menu': 'testimoni',
    }
    return render(request, 'admin_formtestimoni.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def deletetestimoni_admin(request, pk):
    hapus = Testimoni.objects.get(id=pk)
    if request.method == 'POST':
        hapus.delete()
        return redirect('testimoni_admin')

    context = {
        'judul': 'Form Hapus Testimoni',
        'hapus':hapus,
        'menu': 'testimoni',
    }
    return render(request, 'admin_hapustestimoni.html', context)


@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def kontak_admin(request):
    kontak = Kontak.objects.all()
    context = {
        'data': kontak,
        'judul': 'Halaman Kontak',
        'menu': 'kontak',
        
    }
    return render(request, 'admin_kontak.html', context)

@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def laporan(request):
    context = {
    
        'judul': 'Halaman Laporan',
        'menu': 'laporan',
        
    }
    

    if request.method == 'POST':
        jenis_laporan = request.POST.get('jenis_laporan')
        
        if jenis_laporan == 'harian':
            tanggal_dari = datetime.datetime.strptime(request.POST.get('tanggal_dari'), '%Y-%m-%d').date()
            tanggal_sampai = datetime.datetime.strptime(request.POST.get('tanggal_sampai'), '%Y-%m-%d').date()

            # Query transaksi berdasarkan rentang tanggal
            transaksi_list = Transaksi.objects.filter(date_created__range=[tanggal_dari, tanggal_sampai])

        elif jenis_laporan == 'bulanan':
            bulan = int(request.POST.get('bulan'))

            # Query transaksi berdasarkan bulan
            transaksi_list = Transaksi.objects.filter(date_created__month=bulan)

        elif jenis_laporan == 'tahunan':
            tahun = int(request.POST.get('tahun'))

            # Query transaksi berdasarkan tahun
            transaksi_list = Transaksi.objects.filter(date_created__year=tahun)
            
        else:
            transaksi_list = Transaksi.objects.all()
        
        # Membuat file Excel
        workbook = Workbook()
        sheet = workbook.active
        
        # Menulis header
        sheet.cell(row=1, column=1, value='No. Transaksi')
        sheet.cell(row=1, column=2, value='Tanggal')
        sheet.cell(row=1, column=3, value='Alamat Kirim')
        sheet.cell(row=1, column=4, value='Total Transaksi')
        
        # Menulis data transaksi
        row = 2
        for transaksi in transaksi_list:
            sheet.cell(row=row, column=1, value=transaksi.no_transaksi)
            sheet.cell(row=row, column=2, value=transaksi.date_created.strftime('%d-%m-%Y'))
            sheet.cell(row=row, column=3, value=transaksi.alamat_kirim)
            sheet.cell(row=row, column=4, value=transaksi.total_transaksi)
            row += 1
        
        # Mengatur lebar kolom
        column_widths = [15, 12, 20, 15]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width
        
        # Mengatur alignment
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Mengatur nama file
        filename = 'laporan_transaksi.xlsx'
        
        # Mengirim file Excel sebagai response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        
        return response
    return render(request, 'laporan.html',context)
    

@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def transaksi_list(request):
    transaksi = Transaksi.objects.all()
    context = {
        'data': transaksi,
        'judul': 'Halaman Transaksi',
        'menu': 'transaksi',
        
    }
    return render(request, 'admin_transaksi.html', context)
@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def detail_transaksi(request, notransaksi):
    detail = DetailTransaksi.objects.filter(no_transaksi=notransaksi)
    return render(request, 'detail_transaksi.html', {'detail': detail})



@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def custumer_admin(request):
    custumer = Custumer.objects.all()
    context = {
        'data': custumer,
        'judul': 'Halaman Custumer',
        'menu': 'custumer'
    }
    return render(request, 'admin_custumer.html', context)



@login_required(login_url='loginpage')
@ijinkan_pengguna(yang_diizinkan=['administrator']) 
def deletecustumer_admin(request, pk):
    hapus = Custumer.objects.get(id=pk)
    if request.method == 'POST':
        hapus.delete()
        return redirect('custumer_admin')

    context = {
        'judul': 'Form Hapus Custumer',
        'hapus':hapus,
        'menu': 'custumer',
    }
    return render(request, 'admin_hapuscustumer.html', context)


